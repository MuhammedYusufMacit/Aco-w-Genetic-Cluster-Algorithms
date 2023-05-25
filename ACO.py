import math
import random
import pandas as pd
from matplotlib import pyplot as plt
import time
import sympy #çokgenler arasında en kısa iki nokta
import csv
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import numpy as np
from scipy.spatial import distance
from shapely.ops import nearest_points
from shapely.geometry import Polygon
from collections import defaultdict
from sklearn.preprocessing import MinMaxScaler
from openpyxl import load_workbook
import multiprocessing

# FONKSİYONLAR
# weighted_random_choice()
# _select_node()
# tour_construction()
# get_instant_distance()
# get_euclid_distance()
# _add_pheromone()
# _evaporation()
# _aco()
# run()
# plot()
# find_closest_distance_between_polys()
# preprocess4run()
class Node:
    def __init__(self, x_y, index): 
        self.index=index
        self.x_y=x_y

class Bridge_Node:

    def __init__(self, a, d, x_y):
        self.arrival_cluster=a
        self.departure_cluster=d
        self.x_y=x_y

# DEĞİŞKENLER
paths= 'C:\\Users\\TRON PCH\\Documents\\berlin52.xls', 'C:\\Users\\Turtle\\Datasets\\berlin52.xls', 'C:\\Users\\LENOVO\\OneDrive\\Masaüstü\\berlin52.xls'
datasets="berlin52","bier127","ts255","lin318","pr439","pcb442","pr1002"
elbow_n_clusters=[4,1,1,1,1,1]
#"5533343333443454"

path = paths[1]
n_clusters = 4
finishvariable=0.0
totalLength=0.0
rho = 0.5
number_of_iterations=150
colony_size=500
initial_pheromone = 1.0
initial_pheromone_weight = 1.0
cluster_end_points = []
cluster_start_points = []
I_am_here=np.empty([0,2])
b = (0,0)
I_am_here = np.vstack([I_am_here, b])

def preprocess_node(node):
    return preprocess4run(node['x_y'], node['index'], cluster_start_points[(node['index'] - 1) % n_clusters])

def tournament_selection(data):
    if len(data) == 0:
        raise ValueError("No data provided.")

    k = min(3, len(data))
    keys = list(data.keys())
    selected_key = None
    for i in range(k):
        # Choose a random key.
        key = random.choice(keys)
        if selected_key is None or data[key] > data[selected_key]:
            selected_key = key

    return selected_key

def weighted_random_choice(choices):
    max = sum(choices.values())
    pick = random.uniform(0, max)
    current = 0
    for key, value in choices.items():
        current += value
        if current > pick:
            return key

# karıncanın gideceği sıraki düğüm belirleniyor
def _select_node(tour_nodes):
    alpha = 1.0
    beta = 5.0

    # ziyaret edilmiş ve edilmemiş düğümleri belirlemek için
    unvisited_nodes = []
    for node in range(number_of_nodes):
        if node not in tour_nodes:
            unvisited_nodes.append(node)

    # düğüm seçimi için olasılık formülünün paydasını hesaplıyoruz
    probability_denominator = 0.0
    for unvisited_node in unvisited_nodes:
        probability_denominator += (1 / pow(cost_distance[tour_nodes[-1]][unvisited_node], beta)) * \
                                    pow(pheromone[tour_nodes[-1]][unvisited_node], alpha)

    # tüm düğümler için olasılık hesabı yaptık
    probabilities = {}
    for unvisited_node in unvisited_nodes:
        try:
            probabilities[unvisited_node] = pow(pheromone[tour_nodes[-1]][unvisited_node],
                                                alpha) * \
                                            (1 / pow(cost_distance[tour_nodes[-1]][unvisited_node],
                                                        beta)) / probability_denominator
        except ValueError:
            pass  # do nothing

    # en yüksek olasılığa göre gidilecek düğüm seçildi, rulet secimi
    #selected = weighted_random_choice(probabilities)
    #Turnuva seçilim metodu
    selected = tournament_selection(probabilities)
    return selected

# karıncanın bir sonraki gideceğim düğüm belirlendi, burada tüm rota oluşturuluyor
def tour_construction(first_node,nodes,index):
    if first_node==-1:
        #print("Başlangıç Node'u Null geldi, değer 0 olarak atandı")
        first_node=[[0]]
        
    tour_nodes = [first_node]
   
    while len(tour_nodes) < number_of_nodes: 
        ekle = _select_node(tour_nodes)
        tour_nodes.append(ekle)
        
    #index[tour_nodes-1]
    return tour_nodes

# tüm rotası belirlenen karıncanın yolunun ne kadar olduğu belirleniyor
def get_instant_distance(tour_nodes):
    # karıncanın o ana kadar dolaştığı yolun uzunluğunu belirlemek için kullanılıyor
    distance = 0.0
    for i in range(number_of_nodes):
        distance = distance + cost_distance[tour_nodes[i]][
            tour_nodes[(i + 1) % number_of_nodes]]

    return distance


# euclid uzaklığı hesaplandı
def get_euclid_distance(initial_pheromone, number_of_nodes, nodes):
    # uzaklık hesabı yapıp self.cost içerisine kayıt edeceğim için ilk değer atamasına ihtiyacım var
    cost = [[None] * number_of_nodes for _ in range(number_of_nodes)]

    for i in range(number_of_nodes):
        for j in range(i + 1, number_of_nodes):
            # cost için her nokta arasındaki euclid hesabı yapılıyor
            cost[i][j] = cost[j][i] = math.sqrt(
                pow(nodes[i][0] - nodes[j][0], 2.0) + pow(nodes[i][1] - nodes[j][1], 2.0))

    return cost

def _add_pheromone(tour_nodes, distance, cost=1.0):
    delta = initial_pheromone_weight / distance  # delta = 1/C^s formülünden
    for i in range(number_of_nodes):
        pheromone[tour_nodes[i]][tour_nodes[(i + 1) % number_of_nodes]] += cost * delta

    return pheromone

# buharlaşma için
def _evaporation():
    for i in range(number_of_nodes):
        for j in range(i + 1, number_of_nodes):
            pheromone[i][j] *= (1.0 - rho)

    return pheromone

def _aco(nodes,index,first_node):
    # son durumda en iyi düğüm ve uzunluk bilgisini tutmak için kullanılacaklar

    final_best_nodes = None
    final_best_distance = float("inf")

    ants_nodes = [] # [0] * colony_size
    ants_distance = [] # [0] * colony_size

    start = time.perf_counter()

    for step in range(number_of_iterations):
        
        # start1 = time.perf_counter()
        # her iterasyonda karıncalar için rota oluşturulup yol hesabı yapılıyor
        for i in range(colony_size):
            ants_nodes.append(tour_construction(first_node,nodes,index))
        # finish1 = time.perf_counter()
        # print(f'Finished in {round(finish1 - start1, 4)} sec(s)')
        for ant in ants_nodes:
            ants_distance.append(get_instant_distance(ant))

        # oluşturulan rotalardan hangisinin daha iyi olduğu kontrol ediliyor
        for i in range(colony_size):
            pheromone = _add_pheromone(ants_nodes[i], ants_distance[i])
            if ants_distance[i] < final_best_distance:
                final_best_nodes = ants_nodes[i]
                final_best_distance = ants_distance[i]
        # işlem sonucu buharlaşma yapılıyor
        pheromone = _evaporation()
        #plot(nodes, final_best_nodes, mode, labels) #!!!!! 
    
    current_index =[]
    for i in range(len(final_best_nodes)):
        current_index.append(index[final_best_nodes[i]])
            
    finish = time.perf_counter()
    _aco.finishvariable=round(finish - start, 4)
    print(f'Finished in {round(finish - start, 4)} sec(s)')

    return final_best_nodes, final_best_distance, current_index


def run(mode, nodes,index,first_node):
    print('Started : {0}'.format(mode))
    final_best_nodes, final_best_distance, current_index = _aco(nodes,index,first_node)
    print('Ended : {0}'.format(mode))
    print('Sequence : <- {0} ->'.format(' - '.join(str(current_index[i]) for i in range(len(current_index)))))
    #print("TRUE DISTANCE")
    print('Total distance travelled to complete the tour : {0}\n'.format(round(final_best_distance, 2)))
    run.totalLength=round(final_best_distance, 2) 
    return current_index

def calculate_distance(starting_x, starting_y, destination_x, destination_y):
    distance = math.hypot(destination_x - starting_x, destination_y - starting_y)  # calculates Euclidean distance (straight-line) distance between two points
    return distance

def calculate_path(selected_map):
    total_distance = 0
    current_point = selected_map[0]
    for next_point in selected_map[1:]:
        current_distance = calculate_distance(
            current_point[0], current_point[1],
            next_point[0], next_point[1]
        )
        #print(current_point, 'to', next_point, '=', current_distance)
        total_distance += current_distance
        current_point = next_point
    return total_distance

def finalplot(nodes,final_best_nodes):
    
    number_of_nodes=len(nodes)
    labels = range(1, number_of_nodes + 1)
    final_best_nodes = final_best_nodes
    final_best_distance = calculate_path(nodes)
    
    run.totalLength=round(final_best_distance, 2)
    print('Sequence : <- {0} ->'.format(' - '.join(str(final_best_nodes[i]) for i in range(len(final_best_nodes)))))
    print('Total distance travelled to complete the tour : {0}\n'.format(round(final_best_distance, 3)))
    plot(nodes, final_best_nodes, mode, final_best_nodes)
    return final_best_distance


# ekrana grafik çıkarmak için
def plot(nodes, final_best_nodes, mode, labels, line_width=1, point_radius=math.sqrt(2.0), annotation_size=8, dpi=120, save=True, name=None):
    x = []
    y = []
    for i in range(len(final_best_nodes)):
        x.append(nodes[i][0])
        y.append(nodes[i][1])
    x.append(nodes[0][0])
    y.append(nodes[0][1])
    
    plt.plot(x, y, linewidth=line_width)
    plt.scatter(x, y, s=math.pi * (point_radius ** 2.0))
    printvariable="ACO TIME: " + str(_aco.finishvariable) +" ROUTE: "+ str(run.totalLength)+" Iter:" + str(number_of_iterations) + " CSize: " + str(colony_size)
    plt.title(printvariable)
    for i in range(len(final_best_nodes)):
        plt.annotate(int(labels[i]), nodes[i], size=annotation_size)
    if save:
        if name is None:
            name = '{0}.png'.format(mode) 
        plt.savefig(name, dpi=dpi)
    plt.show()
    plt.gcf().clear()


def closest_points(matrix1, matrix2):
    min_distance = math.inf
    closest_points = None
 
    for point1 in matrix1['x_y']:
        for point2 in matrix2['x_y']:
            distance = math.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)
            if distance < min_distance:
                min_distance = distance
                closest_points = (point1, point2)

    return closest_points

def find_closest_distance_between_polys(polygons):

    for i in range(n_clusters):
        for j in range(n_clusters):
            if i == j:
                continue
            nearestpoints = closest_points(polygons[j], polygons[i])
            point_j_to_i=Bridge_Node(j, i,[nearestpoints[0]])
            
            if j-i ==1 or (i==n_clusters-1 and j==0):
                cluster_start_points.append(point_j_to_i)
                
    

    return nearestpoints,cluster_start_points

def find_cluster_end_points(polygons):

    for i in range(n_clusters):
        for j in range(n_clusters):
            if i == j:
                continue
            nearestpoints = closest_points(polygons[i], polygons[j])
            point_i_to_j=Bridge_Node(i, j,[nearestpoints[0]])
            
            if j-i ==1 or (i==n_clusters-1 and j==0):
                cluster_end_points.append(point_i_to_j)

    return cluster_end_points
    
def preprocess4run(nodes,index,first_node):
    global pheromone
    global labels
    global cost_distance
    global number_of_nodes
    number_of_nodes=len(nodes)
    
    distance_calc_time_start = time.perf_counter()
    cost_distance = get_euclid_distance(initial_pheromone, number_of_nodes, nodes)
    distance_calc_time_finish = time.perf_counter()
    
    
    pheromone = [[initial_pheromone] * len(nodes) for _ in range(len(nodes))]
    labels = range(1, number_of_nodes + 1)
    first_node=find_first_node(nodes,first_node)
    final_best_nodes = run(mode, nodes,index,first_node)
    #plot(nodes, final_best_nodes, mode, labels)
    return final_best_nodes

def find_first_node(nodes,first_node):
    index=-1
    for i, arr in enumerate(nodes):
        if (arr == first_node.x_y).all():
            index = i
    return index

def remove_end_points(nodes,i,cluster_end_point):
    for j, x in enumerate(nodes[i]['x_y']):
        if (x == cluster_end_point.x_y[0]).all():
            del nodes[i]['x_y'][j]
            del nodes[i]['index'][j]
    return nodes
    
def final_nodes_concatinating():
   
    finalnodes = np.empty([0,2])
    new_nodes_excel=[]
    new_nodes_excel = np.array(nodes_excel)
    
    for i in range(n_clusters):
        finalnodes = np.concatenate([
            finalnodes,
            nodes_excel[nodes0n[i]-1],
            np.array([cluster_end_points[i].x_y[0]])
            ])

    
    return finalnodes

def find_starting_cluster(cluster_centers_):
    a=np.empty([0,2])
    cluster_centers_ = cluster_centers_.astype(int)
    
    a=np.append(a, I_am_here)
    a=np.append(a, cluster_centers_)
    
    a=a.reshape(int(len(a)/2),2)
    cluster_centers_distances = get_euclid_distance(initial_pheromone, n_clusters+1,a )
    res = []
    for val in cluster_centers_distances[0]:
        if val != None :
            res.append(val)
    return np.argmin(res)
"""
def silhouette(veri_seti):
    # Veri seti yüklenir veya oluşturulur
    X = veri_seti

    k_range = range(2, 10)
    max_silhouette_score = -1
    optimal_k = -1
    
    for k in k_range:
        kmeans = KMeans(n_clusters=k, random_state=42)
        kmeans.fit(X)
        cluster_labels = kmeans.labels_
        silhouette_avg = silhouette_score(X, cluster_labels)
    
        # Silhouette skoru kontrol edilir ve en yüksek skor güncellenirse optimum küme sayısı güncellenir
        if silhouette_avg > max_silhouette_score:
            max_silhouette_score = silhouette_avg
            optimal_k = k
    print("Number of Cluster: ",optimal_k)
    return optimal_k
"""
if __name__ == '__main__':
    
    start = time.perf_counter()
    mode = 'Standard ACO without 2opt'

    nodes_excel = pd.read_excel(path).values
    number_of_nodes = len(nodes_excel)
    global cost_distance
    global pheromone
    
    #n_clusters=silhouette(nodes_excel)
    clustering_time_start = time.perf_counter()
    kmeans = KMeans(n_clusters, init='k-means++', random_state=0).fit(nodes_excel)


    cluster_centers_=kmeans.cluster_centers_
    #find_starting_cluster(cluster_centers_)
    

    nodes_excel = pd.read_excel(path).values
    
    index_values = np.arange(len(nodes_excel)) + 1
    
    nodes_list = []
    for row, index,label in zip(nodes_excel, index_values,kmeans.labels_):
        x_y = np.array(row[:2], dtype=np.float64)  # Get the x, y coordinates
        node = Node(x_y, index)  # Create a new Nodes object
        node.label = label
        nodes_list.append(node)  # Add the new object to the list
 
    grouped_nodes = {label: {'x_y': [], 'index': []} for label in set(kmeans.labels_)}
    for node in nodes_list:
        grouped_nodes[node.label]['x_y'].append(node.x_y)
        grouped_nodes[node.label]['index'].append(node.index)

    cluster_end_points = find_cluster_end_points(grouped_nodes)
        
    for i in range(n_clusters):
        grouped_nodes = remove_end_points(grouped_nodes,i,cluster_end_points[i]) 

    nearest_points_,cluster_start_points = find_closest_distance_between_polys(grouped_nodes)
    
    clustering_time_finish= time.perf_counter()
    clustering_time=round(clustering_time_finish - clustering_time_start, 2)

    distance_calculation_time_start = time.perf_counter()
    """
    pool = multiprocessing.Pool()

    # Map the function to the list of grouped nodes and execute it in parallel
    preprocessed_nodes = pool.map(preprocess_node, grouped_nodes)

    pool.close()
    pool.join()
    """
    preprocessed_nodes = [preprocess4run(grouped_nodes[i]['x_y'],grouped_nodes[i]['index'],cluster_start_points[(i-1)%n_clusters]) for i in range(n_clusters)]
    distance_calculation_time_finish= time.perf_counter()
    distance_calculation_time=round(distance_calculation_time_finish - distance_calculation_time_start, 2)
    
    
    nodes0n = list(map(np.array, preprocessed_nodes))

    finalnodes = final_nodes_concatinating()
 
    nodes=finalnodes.reshape(len(nodes_excel),2)

    #TODO Flattened_deneme isimlendirme 
    final_best_nodes=[]
    koordinat =[]
    indexn = []
    flattened_deneme = []
    
    for i in range(len(nodes)):
        final_best_nodes.append(i)
        
    for i in ((nodes_list)):
        for j in range(n_clusters):
            if(i.x_y == cluster_end_points[j].x_y[0]).all():
                indexn.append(i.index)
        
    for i in range(n_clusters):
        
        flattened_deneme = np.concatenate([
            flattened_deneme,
            nodes0n[i],
            np.array([indexn[i]])
                                           ])


    flattened = [x for sublist in preprocessed_nodes for x in sublist]

    final_best_nodes_final = final_best_nodes
    Improved_ACO_Distance=finalplot(nodes,flattened_deneme)
    
    finish = time.perf_counter()
    
    print("Total Distance: ",str(round(Improved_ACO_Distance,4)))
    print(f'Total Time {round(finish - start, 2)} saniye.')
    print("Colony size: " + str(colony_size))
    print("No Iter: "+ str(number_of_iterations) )
    print("Cluster Count: "+ str(n_clusters))
    Optimal_Distance=input("Best Solution:")
    print("Error Rate: "+str(round((Improved_ACO_Distance - float(Optimal_Distance))/Improved_ACO_Distance,4)))
    print("Distance_calculation_time: "+str(distance_calculation_time))
    print("Clustering Time: "+str(clustering_time))
    print("Average Tour Time: "+str(distance_calculation_time/n_clusters))
   
       
    # Load the existing Excel file
    filename = 'output.xlsx'
    book = load_workbook(filename)
    
    # Select the sheet to write the data
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    
    output_data = {
        "Total Distance": [str(round(Improved_ACO_Distance,4))],
        "Total Time": [round(finish - start, 2)],
        "Colony size": [colony_size],
        "Number of Iteration": [number_of_iterations],
        "Cluster Count": [n_clusters],
        "Optimal Distance": [Optimal_Distance],
        "Error Rate": [round((Improved_ACO_Distance - float(Optimal_Distance))/Improved_ACO_Distance, 4)],
        "Distance Calculation Time": [distance_calculation_time],
        "Clustering Time": [clustering_time],
        "Average Tour Time": [distance_calculation_time/n_clusters],
        "Selection": ["Tournament"],
        "Dataset": ["Berlin52"]
    }
    
    df = pd.DataFrame(output_data)

    # Write DataFrame to the sheet
    df.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row)
    
    writer.save()
    writer.close()
    
    #Eng yorum satırları
    
